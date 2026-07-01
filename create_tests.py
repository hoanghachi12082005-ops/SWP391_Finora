import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment

def format_ws(ws, tcs):
    # Set sheet as active
    wb = ws.parent
    wb.active = ws

    # Delete other workflow sheets
    sheets_to_delete = ['Workflow 1', 'Workflow 2', 'Workflow 3', 'Test Cases']
    for s_name in sheets_to_delete:
        if s_name in wb.sheetnames:
            del wb[s_name]

    # Delete existing rows
    ws.delete_rows(9, ws.max_row - 8)

    # Styles
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    font = Font(name='Arial', size=10)
    alignment = Alignment(wrap_text=True, vertical='top')

    ws.cell(row=9, column=2).value = ws.title
    ws.cell(row=9, column=2).font = Font(bold=True)

    row_idx = 10
    for tc in tcs:
        for col_idx, val in enumerate(tc, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = font
            cell.border = border
            cell.alignment = alignment
        row_idx += 1

def create_category_test():
    wb = openpyxl.load_workbook('d:\\Thangdev\\SWP\\thang\\docs\\System Test.xlsx')
    ws = wb['Staff Management ']
    ws.title = 'Category Management'
    
    tcs = [
        ["CAT-01", "View category list", "1. Navigate to Category page", "Category list is displayed with default pagination.", "Admin/Manager logged in", "Pending"],
        ["CAT-02", "Search category", "1. Enter keyword\n2. Select status\n3. Click Search", "Filtered list is displayed according to criteria.", "Admin/Manager logged in", "Pending"],
        ["CAT-03", "Add category - Success", "1. Click Add\n2. Enter valid name and info\n3. Click Save", "Success message displayed, new category in list.", "Admin/Manager logged in", "Pending"],
        ["CAT-04", "Add category - Empty Name", "1. Click Add\n2. Leave name empty\n3. Click Save", "Error message 'Tên nhóm hàng không được để trống.' displayed.", "Admin/Manager logged in", "Pending"],
        ["CAT-05", "Add category - Duplicate Name", "1. Click Add\n2. Enter existing name\n3. Click Save", "Error message 'Tên nhóm hàng đã tồn tại.' displayed.", "Admin/Manager logged in", "Pending"],
        ["CAT-06", "Update category - Success", "1. Click Edit on a category\n2. Change description\n3. Click Save", "Success message displayed, changes reflected.", "Admin/Manager logged in", "Pending"],
        ["CAT-07", "Update category - Parent is itself", "1. Click Edit\n2. Select itself as parent\n3. Click Save", "Error message 'Nhóm cha không thể là chính nó.' displayed.", "Admin/Manager logged in", "Pending"],
        ["CAT-08", "Update category - Parent is descendant", "1. Click Edit\n2. Select a descendant as parent\n3. Click Save", "Error message 'Nhóm cha không thể là nhóm con của nhóm hiện tại.' displayed.", "Admin/Manager logged in", "Pending"],
        ["CAT-09", "Delete category - Success", "1. Click Delete on category with no dependencies\n2. Confirm", "Success message displayed, category removed.", "Admin/Manager logged in", "Pending"],
        ["CAT-10", "Delete category - Has dependencies", "1. Click Delete on category with products\n2. Confirm", "Error message indicating dependencies exist.", "Admin/Manager logged in", "Pending"]
    ]
    
    format_ws(ws, tcs)
    wb.save('d:\\Thangdev\\SWP\\thang\\docs\\System Test_Category.xlsx')
    print("Updated System Test_Category.xlsx")

def create_kho_test():
    wb = openpyxl.load_workbook('d:\\Thangdev\\SWP\\thang\\docs\\System Test.xlsx')
    ws = wb['Staff Management ']
    ws.title = 'Kho Management'
    
    tcs = [
        ["KHO-01", "Setup initial warehouse", "1. Go to Inventory\n2. Enter warehouse name and address\n3. Click Save", "Initial warehouse created successfully.", "Branch has no warehouse", "Pending"],
        ["KHO-02", "Update warehouse info", "1. Go to Warehouse settings\n2. Edit info\n3. Click Save", "Warehouse info updated successfully.", "Warehouse exists", "Pending"],
        ["KHO-03", "View stock list", "1. Select warehouse\n2. Enter keyword\n3. Click Search", "Stock list filtered accordingly.", "Warehouse has stock", "Pending"],
        ["KHO-04", "Create transfer request", "1. Select products and qty\n2. Select partner warehouse\n3. Click Create Transfer", "Transfer request created, status PENDING.", "Admin/Owner logged in", "Pending"],
        ["KHO-05", "Confirm transfer request", "1. Select PENDING transfer\n2. Click Confirm", "Status IN_TRANSIT, export/import check tickets generated.", "Admin/Owner logged in", "Pending"],
        ["KHO-06", "Cancel transfer request", "1. Select PENDING transfer\n2. Click Cancel", "Ticket cancelled.", "Admin/Owner logged in", "Pending"],
        ["KHO-07", "Reject transfer request", "1. Select PENDING transfer\n2. Click Reject", "Ticket rejected.", "Admin/Owner logged in", "Pending"],
        ["KHO-08", "Confirm Dispatch", "1. Go to Check tab -> Transfer\n2. Select export ticket\n3. Click Confirm Dispatch", "Dispatch confirmed.", "Export ticket exists", "Pending"],
        ["KHO-09", "Reject Dispatch", "1. Go to Check tab -> Transfer\n2. Select export ticket\n3. Enter note\n4. Click Reject", "Dispatch rejected with note.", "Export ticket exists", "Pending"],
        ["KHO-10", "Confirm Receive with Discrepancy", "1. Select import ticket\n2. Enter actual quantities\n3. Click Confirm", "Items received, discrepancy recorded if any.", "Import ticket exists", "Pending"],
        ["KHO-11", "View transfer details", "1. Go to History tab\n2. Click a ticket", "Ticket details and transactions displayed.", "Completed/Rejected ticket exists", "Pending"]
    ]
    
    format_ws(ws, tcs)
    wb.save('d:\\Thangdev\\SWP\\thang\\docs\\System Test_Kho.xlsx')
    print("Updated System Test_Kho.xlsx")

if __name__ == '__main__':
    create_category_test()
    create_kho_test()
