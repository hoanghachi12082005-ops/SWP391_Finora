import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy

def create_detailed_test(file_name, sheet_title, workflow_name, test_req, scenarios):
    wb = openpyxl.load_workbook('d:\\Thangdev\\SWP\\thang\\docs\\System Test.xlsx')
    
    # We will use 'Workflow 1' as the template since it has the correct header
    if 'Workflow 1' in wb.sheetnames:
        ws_template = wb['Workflow 1']
    else:
        # Fallback if Workflow 1 doesn't exist
        ws_template = wb.active
        
    ws_new = wb.copy_worksheet(ws_template)
    ws_new.title = sheet_title
    
    # Update stats area
    ws_new['C1'] = workflow_name
    ws_new['C2'] = test_req
    
    # Count total TCs
    total_tcs = sum(len(tcs) for name, tcs in scenarios)
    ws_new['C3'] = total_tcs
    
    # Keep rows 1 to 8. Delete everything from row 9 downwards
    ws_new.delete_rows(9, ws_new.max_row - 8 + 1)
    
    # Define styles for scenario rows
    scenario_fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid") # Light cyan/blue
    scenario_font = Font(name="Arial", size=10, bold=True)
    border_style = Side(border_style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    alignment_top = Alignment(vertical="top", wrap_text=True)
    
    row_idx = 9
    for scenario_name, tcs in scenarios:
        # Write scenario row
        ws_new.cell(row=row_idx, column=2, value=scenario_name)
        for col in range(2, 17):
            cell = ws_new.cell(row=row_idx, column=col)
            cell.fill = scenario_fill
            cell.font = scenario_font
            cell.border = border
            if col > 2:
                cell.value = None
        # Merge scenario row cells from column B to P
        ws_new.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=16)
        row_idx += 1
        
        # Write test cases
        for tc in tcs:
            tc_data = [
                tc[0], # B: ID
                tc[1], # C: Description
                tc[2], # D: Procedure
                tc[3], # E: Expected Result
                tc[4], # F: Pre-conditions
                "Pending", # G: Round 1
                "",        # H: Test date
                "",        # I: Tester
                "Pending", # J: Round 2
                "",        # K: Test date
                "",        # L: Tester
                "Pending", # M: Round 3
                "",        # N: Test date
                "",        # O: Tester
                ""         # P: Note
            ]
            for i, val in enumerate(tc_data):
                col_idx = i + 2
                cell = ws_new.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.border = border
                cell.alignment = alignment_top
                cell.font = Font(name="Arial", size=10)
            row_idx += 1
            
    # Remove irrelevant sheets
    for s_name in wb.sheetnames:
        if s_name not in ['Cover', 'Test Statistics', sheet_title]:
            del wb[s_name]
            
    wb.active = ws_new
    wb.save('d:\\Thangdev\\SWP\\thang\\docs\\' + file_name)

if __name__ == '__main__':
    category_scenarios = [
        ("Scenario A: Manage Categories List", [
            ["CAT-01", "View category list", "1. Navigate to Category page", "Category list is displayed with default pagination.", "Admin/Manager logged in"],
            ["CAT-02", "Search category", "1. Enter keyword\n2. Select status\n3. Click Search", "Filtered list is displayed according to criteria.", "Admin/Manager logged in"]
        ]),
        ("Scenario B: Add Category", [
            ["CAT-03", "Add category - Success", "1. Click Add\n2. Enter valid name and info\n3. Click Save", "Success message displayed, new category in list.", "Admin/Manager logged in"],
            ["CAT-04", "Add category - Empty Name", "1. Click Add\n2. Leave name empty\n3. Click Save", "Error message 'Tên nhóm hàng không được để trống.' displayed.", "Admin/Manager logged in"],
            ["CAT-05", "Add category - Duplicate Name", "1. Click Add\n2. Enter existing name\n3. Click Save", "Error message 'Tên nhóm hàng đã tồn tại.' displayed.", "Admin/Manager logged in"]
        ]),
        ("Scenario C: Update Category", [
            ["CAT-06", "Update category - Success", "1. Click Edit on a category\n2. Change description\n3. Click Save", "Success message displayed, changes reflected.", "Admin/Manager logged in"],
            ["CAT-07", "Update category - Parent is itself", "1. Click Edit\n2. Select itself as parent\n3. Click Save", "Error message 'Nhóm cha không thể là chính nó.' displayed.", "Admin/Manager logged in"],
            ["CAT-08", "Update category - Parent is descendant", "1. Click Edit\n2. Select a descendant as parent\n3. Click Save", "Error message 'Nhóm cha không thể là nhóm con của nhóm hiện tại.' displayed.", "Admin/Manager logged in"]
        ]),
        ("Scenario D: Delete Category", [
            ["CAT-09", "Delete category - Success", "1. Click Delete on category with no dependencies\n2. Confirm", "Success message displayed, category removed.", "Admin/Manager logged in"],
            ["CAT-10", "Delete category - Has dependencies", "1. Click Delete on category with products\n2. Confirm", "Error message indicating dependencies exist.", "Admin/Manager logged in"]
        ])
    ]
    create_detailed_test("System Test_Category_Detailed.xlsx", "Category Management", "Category Management", "Test all functionalities related to Categories", category_scenarios)
    print("Created System Test_Category_Detailed.xlsx")

    kho_scenarios = [
        ("Scenario A: Manage Warehouse Info", [
            ["KHO-01", "Setup initial warehouse", "1. Go to Inventory\n2. Enter warehouse name and address\n3. Click Save", "Initial warehouse created successfully.", "Branch has no warehouse"],
            ["KHO-02", "Update warehouse info", "1. Go to Warehouse settings\n2. Edit info\n3. Click Save", "Warehouse info updated successfully.", "Warehouse exists"]
        ]),
        ("Scenario B: View Stock List", [
            ["KHO-03", "View stock list", "1. Select warehouse\n2. Enter keyword\n3. Click Search", "Stock list filtered accordingly.", "Warehouse has stock"]
        ]),
        ("Scenario C: Transfer Requests", [
            ["KHO-04", "Create transfer request", "1. Select products and qty\n2. Select partner warehouse\n3. Click Create Transfer", "Transfer request created, status PENDING.", "Admin/Owner logged in"],
            ["KHO-05", "Confirm transfer request", "1. Select PENDING transfer\n2. Click Confirm", "Status IN_TRANSIT, export/import check tickets generated.", "Admin/Owner logged in"],
            ["KHO-06", "Cancel transfer request", "1. Select PENDING transfer\n2. Click Cancel", "Ticket cancelled.", "Admin/Owner logged in"],
            ["KHO-07", "Reject transfer request", "1. Select PENDING transfer\n2. Click Reject", "Ticket rejected.", "Admin/Owner logged in"]
        ]),
        ("Scenario D: Dispatch & Receive Execution", [
            ["KHO-08", "Confirm Dispatch", "1. Go to Check tab -> Transfer\n2. Select export ticket\n3. Click Confirm Dispatch", "Dispatch confirmed.", "Export ticket exists"],
            ["KHO-09", "Reject Dispatch", "1. Go to Check tab -> Transfer\n2. Select export ticket\n3. Enter note\n4. Click Reject", "Dispatch rejected with note.", "Export ticket exists"],
            ["KHO-10", "Confirm Receive with Discrepancy", "1. Select import ticket\n2. Enter actual quantities\n3. Click Confirm", "Items received, discrepancy recorded if any.", "Import ticket exists"],
            ["KHO-11", "View transfer details", "1. Go to History tab\n2. Click a ticket", "Ticket details and transactions displayed.", "Completed/Rejected ticket exists"]
        ])
    ]
    create_detailed_test("System Test_Kho_Detailed.xlsx", "Kho Management", "Warehouse/Inventory Management", "Test all functionalities related to Kho (Inventory)", kho_scenarios)
    print("Created System Test_Kho_Detailed.xlsx")
